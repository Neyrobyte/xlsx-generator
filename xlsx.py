import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import json
import os

# Стандартные стили с исправленными границами
DEFAULT_STYLES = {
    "table": {
        "border": {"all": "thin", "external": "thick"},
        "indent": 1
    },
    "header_row": {
        "color": "D3D3D3",
        "font": {"bold": True},
        "alignment": {"horizontal": "center", "vertical": "center"},
        "wrap_text": True,
        "border": {"external": "thick"}
    },
    "first_column": {
        "color": "F0F0F0",
        "font": {"bold": True},
        "alignment": {"horizontal": "center", "vertical": "center"},
        "wrap_text": True,
        "border": {"external": "thick"}
    },
    "numeric_columns": {
        "even": {"color": "E6F1FF"},  # бледно-голубой
        "odd": {"color": "E6E6FF"}     # бледно-синий
    },
    "text_columns": {
        "even": {"color": "E6FFE6"},   # бледно-зеленый
        "odd": {"color": "FFF2E6"}      # бледно-оранжевый
    },
    "date_columns": {
        "color": "F0E6FF"              # бледно-фиолетовый
    },
    "columns": {
        "border": {"external": "thick", "internal": "thin"}
    }
}

def parse_input_file(filename):
    with open(filename, 'r', encoding='utf-8') as file:
        content = file.read().strip().split('\n\n')
    
    sheets = []
    for sheet_data in content:
        sheet_info = {"name": "Sheet1", "data": []}
        lines = sheet_data.split('\n')
        for line in lines:
            if line.startswith("sheet:"):
                sheet_info["name"] = line.split(":")[1].strip()
            else:
                sheet_info["data"].append(line)
        sheets.append(sheet_info)
    return sheets

def is_number(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

def is_date(value):
    date_formats = [
        r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
        r'\d{2}/\d{2}/\d{4}',   # MM/DD/YYYY
        r'\d{2}\.\d{2}\.\d{4}'  # DD.MM.YYYY
    ]
    for fmt in date_formats:
        if re.fullmatch(fmt, value):
            return True
    return False

def create_border(border_style):
    """Создает границу с указанным стилем для всех сторон"""
    side = Side(border_style=border_style, color="000000")
    return Border(left=side, right=side, top=side, bottom=side)

def apply_default_styles(ws, data):
    # Создаем стили границ
    thin_border = create_border("thin")
    thick_border = create_border("thick")
    
    # Применяем стили ко всей таблице
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(indent=1)
    
    # Устанавливаем толстые внешние границы для всей таблицы
    for row in ws.iter_rows():
        for cell in row:
            # Левый край
            if cell.column == 1:
                cell.border = Border(left=Side(style="thick"), 
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom)
            # Правый край
            if cell.column == ws.max_column:
                cell.border = Border(left=cell.border.left,
                                    right=Side(style="thick"),
                                    top=cell.border.top,
                                    bottom=cell.border.bottom)
            # Верхний край
            if cell.row == 1:
                cell.border = Border(left=cell.border.left,
                                    right=cell.border.right,
                                    top=Side(style="thick"),
                                    bottom=cell.border.bottom)
            # Нижний край
            if cell.row == ws.max_row:
                cell.border = Border(left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=Side(style="thick"))
    
    # Стиль для верхней строки
    if data:
        header_fill = PatternFill(start_color=DEFAULT_STYLES["header_row"]["color"],
                                end_color=DEFAULT_STYLES["header_row"]["color"],
                                fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            # Утолщаем все границы для заголовков
            cell.border = thick_border
    
    # Стиль для первого столбца
    first_col_fill = PatternFill(start_color=DEFAULT_STYLES["first_column"]["color"],
                                end_color=DEFAULT_STYLES["first_column"]["color"],
                                fill_type="solid")
    first_col_font = Font(bold=True)
    first_col_alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    
    for cell in ws.iter_cols(min_col=1, max_col=1):
        for c in cell:
            c.fill = first_col_fill
            c.font = first_col_font
            c.alignment = first_col_alignment
            # Утолщаем все границы для первого столбца
            c.border = thick_border
    
    # Стили для числовых, текстовых и датовых столбцов
    for col_idx, column in enumerate(ws.iter_cols(min_row=2), start=1):
        is_numeric_col = all(is_number(str(cell.value)) for cell in column if cell.value is not None)
        is_date_col = all(is_date(str(cell.value)) for cell in column if cell.value is not None)
        
        if is_date_col:
            fill_color = DEFAULT_STYLES["date_columns"]["color"]
            for cell in column:
                if cell.value:
                    try:
                        cell.value = datetime.strptime(cell.value, '%Y-%m-%d')
                    except ValueError:
                        try:
                            cell.value = datetime.strptime(cell.value, '%d/%m/%Y')
                        except ValueError:
                            try:
                                cell.value = datetime.strptime(cell.value, '%d.%m.%Y')
                            except ValueError:
                                pass
                    cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
        elif is_numeric_col:
            fill_color = (DEFAULT_STYLES["numeric_columns"]["even"]["color"] 
                          if col_idx % 2 == 0 
                          else DEFAULT_STYLES["numeric_columns"]["odd"]["color"])
            for cell in column:
                if cell.value:
                    cell.value = float(cell.value)
        else:
            fill_color = (DEFAULT_STYLES["text_columns"]["even"]["color"] 
                          if col_idx % 2 == 0 
                          else DEFAULT_STYLES["text_columns"]["odd"]["color"])
        
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for cell in column:
            cell.fill = fill
            # Сохраняем существующие границы, добавляя только заливку
            current_border = cell.border
            cell.border = Border(left=current_border.left,
                                right=current_border.right,
                                top=current_border.top,
                                bottom=current_border.bottom)
    
    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

def process_cell_data(cell_data):
    if cell_data.startswith("{") and cell_data.endswith("}"):
        try:
            cell_info = json.loads(cell_data)
            value = cell_info.get("value", "")
            
            if isinstance(value, str) and value.startswith("="):
                return value
            
            if isinstance(value, str) and is_date(value):
                return value
            
            return str(value)
        except json.JSONDecodeError:
            return cell_data
    return cell_data

def generate_xlsx(input_file, output_file):
    sheets_data = parse_input_file(input_file)
    
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for sheet_data in sheets_data:
        ws = wb.create_sheet(title=sheet_data["name"])
        data = []
        
        for line in sheet_data["data"]:
            if line.startswith("sheet:"):
                continue
            
            row_data = []
            for cell_data in line.split(';'):
                cell_data = cell_data.strip()
                processed_data = process_cell_data(cell_data)
                row_data.append(processed_data)
            data.append(row_data)
        
        for row in data:
            ws.append(row)
        
        apply_default_styles(ws, data)
    
    wb.save(output_file)

if __name__ == "__main__":
    input_filename = "xlsx_generate.txt"
    output_filename = "storage/shared/download/output.xlsx"
    
    if not os.path.exists(input_filename):
        print(f"Ошибка: файл {input_filename} не найден!")
        exit(1)
    
    try:
        generate_xlsx(input_filename, output_filename)
        print(f"Файл {output_filename} успешно создан!")
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")